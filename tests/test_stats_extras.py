"""Тесты расширенной статистики: временной ряд, день недели, Excel-экспорт."""

from __future__ import annotations

import io
from datetime import datetime, timedelta

from .conftest import login


def _seed(client, n_days: int = 5) -> None:
    """Создаёт несколько инспекций с разными датами прямо в БД."""
    from app.database import SessionLocal
    from app.models import Inspection, InspectionStatus, User

    with SessionLocal() as db:
        op = db.query(User).filter_by(username="operator1").first()
        base = datetime(2026, 3, 2, 10, 0, 0)  # понедельник
        for i in range(n_days):
            for _ in range(i + 1):
                ins = Inspection(
                    operator_id=op.id,
                    original_filename=f"f{i}.jpg",
                    original_path=f"originals/f{i}.jpg",
                    result_path=None,
                    image_width=1024,
                    image_height=768,
                    status=InspectionStatus.SUCCESS,
                    defects_count=i,
                    avg_confidence=0.5,
                    inference_time_ms=12.0,
                    conf_threshold=0.25,
                )
                ins.created_at = base + timedelta(days=i, hours=i)
                db.add(ins)
        db.commit()


def test_stats_has_daily_and_weekday_buckets(client):
    _seed(client, n_days=5)
    admin = login(client, "admin", "admin12345")
    r = client.get(
        "/api/stats?from_date=2026-03-01T00:00:00&to_date=2026-03-10T00:00:00",
        headers={"Authorization": f"Bearer {admin}"},
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["total_inspections"] == sum(range(1, 6))
    assert len(data["by_day"]) == 10  # 2026-03-01 .. 2026-03-10
    # 2026-03-02 — понедельник, суммарно 1 инспекция.
    by_day = {d["date"]: d for d in data["by_day"]}
    assert by_day["2026-03-02"]["inspections"] == 1
    assert by_day["2026-03-06"]["inspections"] == 5

    wd = {w["weekday"]: w for w in data["by_weekday"]}
    assert wd[0]["weekday_name"] == "Пн"
    # В выборке: пн (1), вт (2), ср (3), чт (4), пт (5)
    assert wd[0]["inspections"] == 1
    assert wd[4]["inspections"] == 5


def test_stats_xlsx_export(client):
    _seed(client, n_days=3)
    admin = login(client, "admin", "admin12345")
    r = client.get(
        "/api/stats/export.xlsx",
        headers={"Authorization": f"Bearer {admin}"},
    )
    assert r.status_code == 200
    # xlsx — это ZIP-архив с сигнатурой "PK".
    assert r.content[:2] == b"PK"
    assert "spreadsheetml" in r.headers["content-type"]

    # Проверяем содержимое книги.
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(r.content))
    assert "Сводка" in wb.sheetnames
    assert "По дням" in wb.sheetnames
    assert "По дню недели" in wb.sheetnames
    assert "По классам" in wb.sheetnames
    assert "По операторам" in wb.sheetnames

    ws = wb["По дням"]
    # Заголовок + хотя бы одна строка данных.
    assert ws.cell(row=1, column=1).value == "Дата"
    assert ws.cell(row=2, column=2).value is not None


def test_stats_requires_manager(client):
    op = login(client, "operator1", "operator12345")
    r = client.get("/api/stats", headers={"Authorization": f"Bearer {op}"})
    assert r.status_code == 403
    r = client.get("/api/stats/export.xlsx", headers={"Authorization": f"Bearer {op}"})
    assert r.status_code == 403


def test_stats_manager_can_read(client):
    _seed(client, n_days=2)
    mgr = login(client, "manager1", "manager12345")
    r = client.get(
        "/api/stats?from_date=2026-03-01T00:00:00&to_date=2026-03-10T00:00:00",
        headers={"Authorization": f"Bearer {mgr}"},
    )
    assert r.status_code == 200, r.text
    assert r.json()["total_inspections"] >= 1
