"""Сводная статистика для руководителя (ТЗ Ф9).

Эндпоинты:

* ``GET /api/stats``           — агрегаты + временной ряд по дням + по дню
                                  недели + разбивка по классам и операторам.
* ``GET /api/stats/export.xlsx`` — выгрузка той же статистики в формате
                                  Excel (.xlsx) для презентации/печати.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from sqlalchemy import func, select
from sqlalchemy.orm import Session


log = logging.getLogger("aoi.stats")

from ..database import get_db
from ..models import Defect, Inspection, InspectionStatus, User
from ..schemas import (
    DailyStat,
    DefectClassStat,
    InspectionStatsOut,
    OperatorStats,
    WeekdayStat,
)
from ..services.detector import NAME_BY_CODE
from .deps import require_manager


router = APIRouter(prefix="/api/stats", tags=["Статистика"])


WEEKDAY_RU = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]


def _collect_stats(
    db: Session,
    *,
    from_date: datetime | None,
    to_date: datetime | None,
) -> InspectionStatsOut:
    base_filter = [Inspection.status == InspectionStatus.SUCCESS]
    if from_date:
        base_filter.append(Inspection.created_at >= from_date)
    if to_date:
        base_filter.append(Inspection.created_at <= to_date)

    total_inspections = int(
        db.execute(
            select(func.count()).select_from(Inspection).where(*base_filter)
        ).scalar_one()
        or 0
    )
    defective = int(
        db.execute(
            select(func.count())
            .select_from(Inspection)
            .where(*base_filter, Inspection.defects_count > 0)
        ).scalar_one()
        or 0
    )
    total_defects = int(
        db.execute(
            select(func.coalesce(func.sum(Inspection.defects_count), 0)).where(*base_filter)
        ).scalar_one()
        or 0
    )

    class_stmt = (
        select(Defect.class_code, func.count(Defect.id))
        .join(Inspection, Inspection.id == Defect.inspection_id)
        .where(*base_filter)
        .group_by(Defect.class_code)
        .order_by(func.count(Defect.id).desc())
    )
    by_class = [
        DefectClassStat(
            class_code=code,
            class_name=NAME_BY_CODE.get(code, code),
            count=int(count),
        )
        for code, count in db.execute(class_stmt).all()
    ]

    op_stmt = (
        select(
            User.id,
            User.username,
            User.full_name,
            func.count(Inspection.id),
            func.coalesce(func.sum(Inspection.defects_count), 0),
        )
        .join(Inspection, Inspection.operator_id == User.id)
        .where(*base_filter)
        .group_by(User.id, User.username, User.full_name)
        .order_by(func.count(Inspection.id).desc())
    )
    by_operator = [
        OperatorStats(
            operator_id=uid,
            username=username,
            full_name=full_name or "",
            inspections_count=int(insp_cnt),
            defects_count=int(def_cnt),
        )
        for uid, username, full_name, insp_cnt, def_cnt in db.execute(op_stmt).all()
    ]

    # Временной ряд: считаем в Python на основе выборки дат инспекций.
    # Так избегаем различий между SQLite и PostgreSQL в функциях date_trunc.
    rows = db.execute(
        select(Inspection.created_at, Inspection.defects_count).where(*base_filter)
    ).all()

    daily: dict[str, list[int]] = {}  # date -> [insp, defects]
    wd: dict[int, list[int]] = {i: [0, 0] for i in range(7)}
    for created_at, defects_count in rows:
        if created_at is None:
            continue
        key = created_at.strftime("%Y-%m-%d")
        bucket = daily.setdefault(key, [0, 0])
        bucket[0] += 1
        bucket[1] += int(defects_count or 0)
        wd[created_at.weekday()][0] += 1
        wd[created_at.weekday()][1] += int(defects_count or 0)

    # Плотный временной ряд (заполняем дыры нулями, чтобы красиво рисовать
    # график). Диапазон берём от from_date до to_date, а если фильтр пуст —
    # от первой до последней инспекции.
    if rows:
        lo = min(r[0] for r in rows if r[0])
        hi = max(r[0] for r in rows if r[0])
    else:
        lo = hi = None
    start = (from_date or lo)
    end = (to_date or hi)
    by_day: list[DailyStat] = []
    if start is not None and end is not None:
        day = datetime(start.year, start.month, start.day)
        last = datetime(end.year, end.month, end.day)
        # Защита от слишком длинного периода на UI.
        max_days = 400
        i = 0
        while day <= last and i < max_days:
            key = day.strftime("%Y-%m-%d")
            insp, defs = daily.get(key, (0, 0))
            by_day.append(DailyStat(date=key, inspections=int(insp), defects=int(defs)))
            day += timedelta(days=1)
            i += 1

    by_weekday: list[WeekdayStat] = [
        WeekdayStat(
            weekday=i,
            weekday_name=WEEKDAY_RU[i],
            inspections=int(wd[i][0]),
            defects=int(wd[i][1]),
        )
        for i in range(7)
    ]

    return InspectionStatsOut(
        total_inspections=total_inspections,
        total_defects=total_defects,
        defective_inspections=defective,
        clean_inspections=total_inspections - defective,
        by_class=by_class,
        by_operator=by_operator,
        by_day=by_day,
        by_weekday=by_weekday,
        from_date=from_date,
        to_date=to_date,
    )


@router.get(
    "",
    response_model=InspectionStatsOut,
    summary="Сводная статистика инспекций (Ф9, только руководитель)",
)
def stats_summary(
    from_date: datetime | None = Query(default=None),
    to_date: datetime | None = Query(default=None),
    _: User = Depends(require_manager),
    db: Session = Depends(get_db),
) -> InspectionStatsOut:
    return _collect_stats(db, from_date=from_date, to_date=to_date)


@router.get(
    "/export.xlsx",
    summary="Экспорт статистики в Excel (.xlsx)",
    response_class=Response,
)
def stats_export_xlsx(
    from_date: datetime | None = Query(default=None),
    to_date: datetime | None = Query(default=None),
    _: User = Depends(require_manager),
    db: Session = Depends(get_db),
) -> Response:
    stats = _collect_stats(db, from_date=from_date, to_date=to_date)

    # openpyxl импортируем лениво: это зависимость, нужная только для этого
    # маршрута. Если её нет — отдаём 503 с понятным текстом, а не глухой 500.
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        log.error("openpyxl не установлен: %s", exc)
        raise HTTPException(
            status_code=503,
            detail=(
                "Пакет openpyxl не установлен на сервере. "
                "Выполните 'pip install openpyxl' и перезапустите сервис."
            ),
        ) from exc

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4ED8")
    center = Alignment(horizontal="center", vertical="center")

    def _write_header(ws, headers: list[str]) -> None:
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        for col, _ in enumerate(headers, start=1):
            ws.column_dimensions[get_column_letter(col)].width = 22

    # --- Лист «Сводка» ---
    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Показатель"
    ws["B1"] = "Значение"
    ws["A1"].font = header_font
    ws["B1"].font = header_font
    ws["A1"].fill = header_fill
    ws["B1"].fill = header_fill
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    summary_rows = [
        ("Период с",               (from_date.strftime("%Y-%m-%d %H:%M") if from_date else "—")),
        ("Период по",              (to_date.strftime("%Y-%m-%d %H:%M")   if to_date   else "—")),
        ("Всего инспекций",        stats.total_inspections),
        ("С дефектами",            stats.defective_inspections),
        ("Чистых",                 stats.clean_inspections),
        ("Всего дефектов",         stats.total_defects),
    ]
    for i, (k, v) in enumerate(summary_rows, start=2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

    # --- Лист «По дням» ---
    ws_d = wb.create_sheet("По дням")
    _write_header(ws_d, ["Дата", "Инспекций", "Дефектов"])
    for i, d in enumerate(stats.by_day, start=2):
        ws_d.cell(row=i, column=1, value=d.date)
        ws_d.cell(row=i, column=2, value=d.inspections)
        ws_d.cell(row=i, column=3, value=d.defects)

    # --- Лист «По дню недели» ---
    ws_w = wb.create_sheet("По дню недели")
    _write_header(ws_w, ["День недели", "Инспекций", "Дефектов"])
    for i, w in enumerate(stats.by_weekday, start=2):
        ws_w.cell(row=i, column=1, value=w.weekday_name)
        ws_w.cell(row=i, column=2, value=w.inspections)
        ws_w.cell(row=i, column=3, value=w.defects)

    # --- Лист «По классам» ---
    ws_c = wb.create_sheet("По классам")
    _write_header(ws_c, ["Код", "Класс дефекта", "Количество"])
    for i, c in enumerate(stats.by_class, start=2):
        ws_c.cell(row=i, column=1, value=c.class_code)
        ws_c.cell(row=i, column=2, value=c.class_name)
        ws_c.cell(row=i, column=3, value=c.count)

    # --- Лист «По операторам» ---
    ws_o = wb.create_sheet("По операторам")
    _write_header(ws_o, ["Логин", "ФИО", "Инспекций", "Дефектов"])
    for i, o in enumerate(stats.by_operator, start=2):
        ws_o.cell(row=i, column=1, value=o.username)
        ws_o.cell(row=i, column=2, value=o.full_name)
        ws_o.cell(row=i, column=3, value=o.inspections_count)
        ws_o.cell(row=i, column=4, value=o.defects_count)

    try:
        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()
    except Exception as exc:  # pragma: no cover — страховка на проде
        log.exception("Ошибка сборки XLSX")
        raise HTTPException(status_code=500, detail=f"Не удалось собрать Excel: {exc}") from exc

    suffix = ""
    if from_date or to_date:
        a = from_date.strftime("%Y%m%d") if from_date else "all"
        b = to_date.strftime("%Y%m%d") if to_date else "all"
        suffix = f"_{a}-{b}"
    filename = f"aoi_stats{suffix}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
